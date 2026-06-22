"""Catalog extraction sits behind a clean interface so the internals
(LLM model, vision model, parser) can change without touching the marketplace.

Two implementations ship in M1:
  * StructuredExtractor - CSV / XLSX catalogs (no LLM; runs fully offline).
  * LLMExtractor        - PDF / HTML catalogs via an LLM (model wiring is the
                          one external dependency; the prompt + JSON-guarded
                          parse are real).
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Protocol

from astor.catalog.schemas import ExtractedProduct
from astor.config import settings

_FIELDS = set(ExtractedProduct.model_fields.keys())


class Extractor(Protocol):
    def extract(self, source: Path) -> list[ExtractedProduct]: ...


def _coerce(row: dict) -> ExtractedProduct:
    clean = {k: (v if v not in ("", None) else None) for k, v in row.items() if k in _FIELDS}
    specs_raw = row.get("specs")
    if isinstance(specs_raw, str) and specs_raw.strip():
        try:
            clean["specs"] = json.loads(specs_raw)
        except json.JSONDecodeError:
            clean["specs"] = {"_raw": specs_raw}
    return ExtractedProduct(**clean)


class StructuredExtractor:
    """For supplier catalogs already in CSV / XLSX (the common first case)."""

    def extract(self, source: Path) -> list[ExtractedProduct]:
        if source.suffix.lower() in {".xlsx", ".xlsm"}:
            return self._from_xlsx(source)
        return self._from_csv(source)

    def _from_csv(self, source: Path) -> list[ExtractedProduct]:
        with source.open(newline="", encoding="utf-8-sig") as fh:
            return [_coerce(row) for row in csv.DictReader(fh)]

    def _from_xlsx(self, source: Path) -> list[ExtractedProduct]:
        from openpyxl import load_workbook

        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        out: list[ExtractedProduct] = []
        for r in rows:
            out.append(_coerce(dict(zip(header, r))))
        return out


class LLMExtractor:
    """For unstructured catalogs (PDF/HTML). Emits the SAME DTO as the structured
    path so downstream code is identical. Output is JSON-guarded against the schema."""

    SYSTEM = (
        "Extract life-science catalog products as a JSON array. Each item must have: "
        "supplier_sku, name, category, brand, mpn, pack_size, cost, currency, stock, "
        "lead_time_days, specs (object). Use null for unknown fields. Output ONLY JSON."
    )

    def __init__(self, model: str = "claude-sonnet-4-6") -> None:
        self.model = model

    def extract(self, source: Path) -> list[ExtractedProduct]:
        if not settings.anthropic_api_key:
            raise RuntimeError(
                "LLMExtractor needs ANTHROPIC_API_KEY. Use StructuredExtractor for CSV/XLSX."
            )
        from anthropic import Anthropic

        client = Anthropic(api_key=settings.anthropic_api_key)
        text = source.read_text(encoding="utf-8", errors="ignore")
        msg = client.messages.create(
            model=self.model,
            max_tokens=4096,
            system=self.SYSTEM,
            messages=[{"role": "user", "content": text[:120_000]}],
        )
        payload = "".join(b.text for b in msg.content if getattr(b, "type", None) == "text")
        payload = payload.strip().removeprefix("```json").removeprefix("```").removesuffix("```")
        data = json.loads(payload)
        # Guardrail: every item must validate against the DTO, or it is dropped.
        out: list[ExtractedProduct] = []
        for item in data:
            try:
                out.append(ExtractedProduct(**item))
            except Exception:
                continue
        return out


def for_source(source: Path) -> Extractor:
    """Pick an extractor by file type."""
    if source.suffix.lower() in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        return StructuredExtractor()
    return LLMExtractor()
